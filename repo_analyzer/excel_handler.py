"""
Excel file operations for repository analysis
"""
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class ExcelHandler:
    """Handle Excel file operations for repository analysis"""

    @staticmethod
    def read_input_file(input_file):
        """
        Read data from input Excel file

        Args:
            input_file: Path to input Excel file with GitHub URLs

        Returns:
            List of dictionaries with repo data
        """
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"Input file '{input_file}' not found.")

        print(f"Reading data from {input_file}...")

        wb = load_workbook(input_file)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Find column indices
        try:
            id_col = headers.index('ID')
            timestamp_col = headers.index('TimeStamp')
            subject_col = headers.index('Subject')
            search_col = headers.index('Search Criteria')
            url_col = headers.index('github Repo URL')
        except ValueError as e:
            raise ValueError(f"Required column not found in Excel file: {e}")

        # Read data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[id_col]:  # Skip empty rows
                data.append({
                    'id': str(row[id_col]),
                    'timestamp': row[timestamp_col],
                    'subject': row[subject_col],
                    'search_criteria': row[search_col],
                    'github_url': row[url_col] if row[url_col] else '',
                    'total_lines': 0,
                    'small_files_lines': 0,
                    'grade': 0.0,
                    'status': 'pending'
                })

        print(f"Found {len(data)} repositories to analyze")
        return data

    @staticmethod
    def export_results(repos_data, output_file, small_file_threshold):
        """
        Export analyzed data to Excel file

        Args:
            repos_data: List of repository data dictionaries
            output_file: Path to output Excel file
            small_file_threshold: Threshold used for small files
        """
        print(f"\nExporting results to {output_file}...")

        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Repo Analysis"

        # Define headers
        headers = [
            'ID',
            'TimeStamp',
            'Subject',
            'Search Criteria',
            'github Repo URL',
            'Total Lines',
            f'Lines in Small Files (<{small_file_threshold})',
            'Grade (%)'
        ]
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add data rows
        for repo in repos_data:
            ws.append([
                repo['id'],
                repo['timestamp'],
                repo['subject'],
                repo['search_criteria'],
                repo['github_url'],
                repo['total_lines'] if repo['status'] == 'analyzed' else 'N/A',
                repo['small_files_lines'] if repo['status'] == 'analyzed' else 'N/A',
                repo['grade'] if repo['status'] == 'analyzed' else 'N/A'
            ])

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15

        # Save the workbook
        wb.save(output_file)
        print(f"✓ Successfully exported to {output_file}")
