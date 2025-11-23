#!/usr/bin/env python3
"""
Message Writer Agent - Generate personalized feedback messages based on code analysis grades
"""
import os
import sys
import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class MessageWriter:
    """Agent to generate personalized feedback messages based on grades"""

    def __init__(self, input_file='Output_23.xlsx', output_file='Output_34.xlsx'):
        """

from logger_config import LoggerConfig

# Setup logger
logger = LoggerConfig.setup_logger('message_writer')

        Initialize the Message Writer

        Args:
            input_file: Path to input Excel file with grades
            output_file: Path to output Excel file with messages
        """
        self.input_file = input_file
        self.output_file = output_file
        self.data = []

    def read_excel_data(self):
        """
        Read data from input Excel file

        Returns:
            List of dictionaries with repo data and grades
        """
        if not os.path.exists(self.input_file):
            raise FileNotFoundError(f"Input file '{self.input_file}' not found.")

        print(f"Reading data from {self.input_file}...")

        wb = load_workbook(self.input_file)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Find column indices
        try:
            id_col = headers.index('ID')
            timestamp_col = headers.index('TimeStamp')
            subject_col = headers.index('Subject')
            search_col = headers.index('Search Criteria')
            url_col = headers.index('github Repo URL')
            total_lines_col = headers.index('Total Lines')
            small_files_col = next(i for i, h in enumerate(headers) if 'Lines in Small Files' in str(h))
            grade_col = next(i for i, h in enumerate(headers) if 'Grade' in str(h))
        except (ValueError, StopIteration) as e:
            raise ValueError(f"Required column not found in Excel file: {e}")

        # Read data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[id_col]:  # Skip empty rows
                # Parse grade value (handle both numeric and string with %)
                grade_val = row[grade_col]
                if isinstance(grade_val, str):
                    grade_val = grade_val.replace('%', '').strip()
                try:
                    grade = float(grade_val) if grade_val not in ['N/A', None, ''] else 0.0
                except (ValueError, TypeError):
                    grade = 0.0

                data.append({
                    'id': str(row[id_col]),
                    'timestamp': row[timestamp_col],
                    'subject': row[subject_col],
                    'search_criteria': row[search_col],
                    'github_url': row[url_col],
                    'total_lines': row[total_lines_col],
                    'small_files_lines': row[small_files_col],
                    'grade': grade,
                    'message': ''
                })

        print(f"Found {len(data)} repositories to process")
        return data

    def generate_trump_message(self, repo_data):
        """
        Generate congratulation message in Donald Trump's style (90+)

        Args:
            repo_data: Dictionary with repository data

        Returns:
            Congratulatory message string
        """
        grade = repo_data['grade']

        messages = [
            f"INCREDIBLE! Absolutely INCREDIBLE! This code is {grade}% modular - that's TREMENDOUS! "
            f"Nobody writes code this good. Nobody. I've seen a lot of code, believe me, and this is "
            f"THE BEST. Beautiful, clean, modular - just perfect. This developer is a WINNER. "
            f"We need more developers like this. The BEST developers. Fantastic job!",

            f"WOW! {grade}% modularity - that's AMAZING! This is what I call WINNING code! "
            f"Very professional, very clean. The files are small, organized - just the way it should be. "
            f"I know quality when I see it, and this is QUALITY. Top tier. First class. "
            f"This developer gets it. They really get it. EXCELLENT work!",

            f"Let me tell you something - this code is SPECTACULAR! {grade}% modular. That's the kind of "
            f"number winners get. BIG LEAGUE coding right here. The structure is perfect, the organization "
            f"is perfect - everything is just PERFECT. This is how you write code. "
            f"Tremendous achievement. Really tremendous. CONGRATULATIONS!",
        ]

        # Rotate through messages based on ID
        idx = hash(repo_data['id']) % len(messages)
        return messages[idx]

    def generate_netanyahu_message(self, repo_data):
        """
        Generate positive feedback in Benjamin Netanyahu's style (70-89)

        Args:
            repo_data: Dictionary with repository data

        Returns:
            Positive feedback message string
        """
        grade = repo_data['grade']

        messages = [
            f"Let me be clear: achieving {grade}% modularity demonstrates solid technical capability. "
            f"The evidence shows a well-structured codebase with thoughtful organization. "
            f"This level of modular design reflects an understanding of best practices and maintainability. "
            f"The data speaks for itself - this is commendable work. With continued focus on these "
            f"principles, even greater achievements lie ahead. Well done.",

            f"Analysis of this codebase reveals {grade}% modularity - a strong result by any measure. "
            f"History teaches us that quality code is built through discipline and attention to structure. "
            f"This developer has demonstrated both. The small file architecture shows strategic thinking "
            f"and commitment to maintainability. This is the foundation upon which robust systems are built. "
            f"Good work. Continue on this path.",

            f"The metrics are clear: {grade}% modularity represents solid engineering. Throughout the "
            f"history of software development, we've learned that modular code leads to sustainable systems. "
            f"This codebase reflects that understanding. The balance between complexity and organization "
            f"is well-managed. This developer has made good choices. The foundation is strong. "
            f"Keep building on this success.",
        ]

        idx = hash(repo_data['id']) % len(messages)
        return messages[idx]

    def generate_hason_message(self, repo_data):
        """
        Generate improvement message in Shahar Hason's style (50-69)

        Args:
            repo_data: Dictionary with repository data

        Returns:
            Humorous improvement message string
        """
        grade = repo_data['grade']

        messages = [
            f"אז... {grade}% modularity. לא רע, לא רע בכלל! (Not bad at all!) "
            f"But listen, we can do better here, right? It's like going to the gym - "
            f"you're doing good, but maybe add a few more reps? 😊 The code is okay, "
            f"some files are nice and small, but there's room to break things down more. "
            f"Think of it like hummus - better in small containers than one huge bucket! "
            f"You're on the right track, my friend. Just needs a bit more organization. "
            f"Keep going! 💪",

            f"So I looked at this code... {grade}% modular. אחלה התחלה! (Great start!) "
            f"You know what this reminds me of? My closet. Some things are organized, "
            f"some things... not so much. 😄 But that's okay! We all have that one drawer "
            f"that's messy. The important thing is you're trying! Break those big files "
            f"down a bit more, make it easier to find things. You got this! "
            f"I believe in you! ✨",

            f"{grade}% modularity - רגע, רגע (wait, wait)... this is like a falafel that's "
            f"good but could be GREAT with more tehina! 😋 The foundation is there, "
            f"the potential is there, but let's make those files smaller, yeah? "
            f"Break it down like you're explaining to your grandma - small pieces, "
            f"easy to understand. You're doing fine! Just needs some love. "
            f"Come on, you can do better! I'm rooting for you! 🎉",
        ]

        idx = hash(repo_data['id']) % len(messages)
        return messages[idx]

    def generate_amsalem_message(self, repo_data):
        """
        Generate brutally honest message in Dudi Amsalem's style (<50)

        Args:
            repo_data: Dictionary with repository data

        Returns:
            Strong, direct feedback message string
        """
        grade = repo_data['grade']

        messages = [
            f"תקשיב טוב (Listen well) - {grade}% modularity?! זה לא מקובל! (This is unacceptable!) "
            f"What is this? Giant files, no organization, everything mixed together like a mess! "
            f"This is exactly the problem - no discipline, no structure! You think this is how "
            f"professionals write code?! די כבר! (Enough already!) Break these files down! "
            f"Small, focused, organized - that's what we need! Not this chaos! "
            f"Fix this immediately. This is not acceptable. We expect MUCH better! 💢",

            f"מה זה?! (What is this?!) {grade}% modular? This is a disaster! "
            f"Big files everywhere, no separation of concerns, no organization! "
            f"האם זה ברצינות?! (Are you serious?!) This is the kind of code that creates problems! "
            f"We need standards! We need quality! Not this mess! עוד פעם נגיד את זה - "
            f"(We'll say it again) - BREAK IT DOWN! Small files! Clear structure! "
            f"This needs MAJOR improvement. Now. לא מחר! (Not tomorrow!) NOW! ⚠️",

            f"אני לא מאמין! (I don't believe it!) {grade}% modularity is UNACCEPTABLE! "
            f"This is sloppy work! Everything in huge files, no thought about maintainability! "
            f"זה בדיוק מה שלא צריך לעשות! (This is exactly what you shouldn't do!) "
            f"You want to be taken seriously? Then write serious code! Organized! Modular! "
            f"Not this חוסר סדר (disorder)! We're not playing games here! "
            f"Fix this code RIGHT NOW! We need to see improvement IMMEDIATELY! 🔥",
        ]

        idx = hash(repo_data['id']) % len(messages)
        return messages[idx]

    def generate_message(self, repo_data):
        """
        Generate appropriate message based on grade

        Args:
            repo_data: Dictionary with repository data

        Returns:
            Generated message string
        """
        grade = repo_data['grade']

        if grade >= 90:
            return self.generate_trump_message(repo_data)
        elif grade >= 70:
            return self.generate_netanyahu_message(repo_data)
        elif grade >= 50:
            return self.generate_hason_message(repo_data)
        else:
            return self.generate_amsalem_message(repo_data)

    def process_all(self):
        """
        Process all repositories and generate messages

        Returns:
            List of processed data with messages
        """
        # Read data
        self.data = self.read_excel_data()

        print(f"\n{'='*70}")
        print(f"Generating Personalized Messages")
        print(f"{'='*70}\n")

        # Generate messages for each repo
        for i, repo in enumerate(self.data, 1):
            grade = repo['grade']
            message = self.generate_message(repo)
            repo['message'] = message

            # Determine style used
            if grade >= 90:
                style = "Trump (Congratulations)"
            elif grade >= 70:
                style = "Netanyahu (Positive)"
            elif grade >= 50:
                style = "Hason (Improvement)"
            else:
                style = "Amsalem (Brutally Honest)"

            print(f"[{i}/{len(self.data)}] {repo['id']} - Grade: {grade:.2f}% - Style: {style}")

        print(f"\n✓ Generated {len(self.data)} personalized messages")
        return self.data

    def export_to_excel(self):
        """
        Export data with messages to Excel file
        """
        if not self.data:
            print("No data to export. Run process_all() first.")
            return

        print(f"\nExporting results to {self.output_file}...")

        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Repo Analysis with Feedback"

        # Define headers
        headers = [
            'ID',
            'TimeStamp',
            'Subject',
            'Search Criteria',
            'github Repo URL',
            'Total Lines',
            'Lines in Small Files',
            'Grade (%)',
            'Feedback Message'
        ]
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add data rows
        for repo in self.data:
            ws.append([
                repo['id'],
                repo['timestamp'],
                repo['subject'],
                repo['search_criteria'],
                repo['github_url'],
                repo['total_lines'],
                repo['small_files_lines'],
                repo['grade'],
                repo['message']
            ])

        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # ID
        ws.column_dimensions['B'].width = 20  # TimeStamp
        ws.column_dimensions['C'].width = 50  # Subject
        ws.column_dimensions['D'].width = 30  # Search Criteria
        ws.column_dimensions['E'].width = 60  # github Repo URL
        ws.column_dimensions['F'].width = 15  # Total Lines
        ws.column_dimensions['G'].width = 25  # Lines in Small Files
        ws.column_dimensions['H'].width = 15  # Grade (%)
        ws.column_dimensions['I'].width = 100  # Feedback Message

        # Enable text wrapping for message column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Save the workbook
        wb.save(self.output_file)
        print(f"✓ Successfully exported to {self.output_file}")

    def run(self):
        """
        Run the complete message writing pipeline

        Returns:
            List of processed data with messages
        """
        try:
            # Process all repositories
            self.process_all()

            # Export to Excel
            self.export_to_excel()

            # Print summary
            self.print_summary()

            return self.data

        except Exception as e:
            print(f"\nError: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)

    def print_summary(self):
        """Print summary of message generation"""
        print(f"\n{'='*70}")
        print(f"Message Generation Summary")
        print(f"{'='*70}")

        trump_count = sum(1 for r in self.data if r['grade'] >= 90)
        netanyahu_count = sum(1 for r in self.data if 70 <= r['grade'] < 90)
        hason_count = sum(1 for r in self.data if 50 <= r['grade'] < 70)
        amsalem_count = sum(1 for r in self.data if r['grade'] < 50)

        print(f"Total Repositories: {len(self.data)}")
        print(f"\nMessages by Style:")
        print(f"  Trump (90-100%):     {trump_count} - Congratulations")
        print(f"  Netanyahu (70-89%):  {netanyahu_count} - Positive Feedback")
        print(f"  Hason (50-69%):      {hason_count} - Needs Improvement")
        print(f"  Amsalem (<50%):      {amsalem_count} - Brutally Honest")


def main():
    """Main function to run the message writer"""
    parser = argparse.ArgumentParser(
        description='Message Writer Agent - Generate personalized feedback based on code grades',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example:
  # Generate messages from Output_23.xlsx
  python message_writer.py --input Output_23.xlsx

  # Specify custom output file
  python message_writer.py --input Output_23.xlsx --output custom_messages.xlsx

Message Styles by Grade:
  90-100%:  Donald Trump style (Congratulations!)
  70-89%:   Benjamin Netanyahu style (Positive feedback)
  50-69%:   Shahar Hason style (Needs improvement, humorous)
  0-49%:    Dudi Amsalem style (Brutally honest, direct)
        """
    )

    parser.add_argument(
        '--input',
        type=str,
        default='Output_23.xlsx',
        help='Input Excel file with grades (default: Output_23.xlsx)'
    )

    parser.add_argument(
        '--output',
        type=str,
        default='Output_34.xlsx',
        help='Output Excel file with messages (default: Output_34.xlsx)'
    )

    args = parser.parse_args()

    # Create and run message writer
    writer = MessageWriter(
        input_file=args.input,
        output_file=args.output
    )

    writer.run()


if __name__ == '__main__':
    main()
