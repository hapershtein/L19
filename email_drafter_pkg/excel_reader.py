"""
Excel reader for email drafting
"""
import os
import sys
import openpyxl
from logger_config import LoggerConfig

logger = LoggerConfig.setup_logger('email_drafter')


class ExcelReader:
    """Read feedback messages from Excel files"""

    @staticmethod
    def read_feedback_data(input_file):
        """
        Read feedback messages from Excel file

        Args:
            input_file: Path to Excel file with feedback messages

        Returns:
            List of dictionaries with feedback data
        """
        logger.info(f"Reading data from {input_file}")
        print(f"Reading data from {input_file}...")

        if not os.path.exists(input_file):
            print(f"Error: Input file '{input_file}' not found!")
            sys.exit(1)

        try:
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]

            required_columns = ['ID', 'Feedback Message']
            column_indices = {}

            for col in required_columns:
                try:
                    column_indices[col] = headers.index(col)
                except ValueError:
                    for idx, header in enumerate(headers):
                        if header and col.lower() in header.lower():
                            column_indices[col] = idx
                            break
                    else:
                        print(f"Error: Required column '{col}' not found in Excel file")
                        print(f"Available columns: {headers}")
                        sys.exit(1)

            email_col = None
            subject_col = None
            for idx, header in enumerate(headers):
                if header:
                    if 'subject' in header.lower():
                        subject_col = idx
                    elif 'email' in header.lower() or 'recipient' in header.lower():
                        email_col = idx

            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[column_indices['ID']]:
                    continue

                repo_id = row[column_indices['ID']]
                feedback_message = row[column_indices['Feedback Message']]

                if not feedback_message or feedback_message == 'N/A':
                    continue

                data_row = {
                    'id': str(repo_id),
                    'feedback': str(feedback_message),
                }

                if subject_col is not None and row[subject_col]:
                    data_row['subject'] = str(row[subject_col])
                if email_col is not None and row[email_col]:
                    data_row['recipient'] = str(row[email_col])

                data.append(data_row)

        except Exception as e:
            print(f"Error reading Excel file: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)

        logger.info(f"Found {len(data)} feedback messages to draft")
        print(f"Found {len(data)} feedback messages to draft\n")
        return data
