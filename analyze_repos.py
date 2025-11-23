#!/usr/bin/env python3
"""
Analyze Repos Agent - Clone GitHub repositories and analyze code metrics
"""
import os
import sys
import asyncio
import argparse
import shutil
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from logger_config import LoggerConfig

# Setup logger
logger = LoggerConfig.setup_logger('analyze_repos')


class RepoAnalyzer:
    """Agent to clone and analyze GitHub repositories"""

    def __init__(self, input_file, output_file='Output_23.xlsx', temp_dir='TempFiles', small_file_threshold=150):
        """
        Initialize the Repo Analyzer

        Args:
            input_file: Path to input Excel file with GitHub URLs
            output_file: Path to output Excel file
            temp_dir: Directory to store cloned repositories
            small_file_threshold: Maximum line count for a file to be considered "small" (default: 150)
        """
        self.input_file = input_file
        self.output_file = output_file
        self.temp_dir = temp_dir
        self.small_file_threshold = small_file_threshold
        self.repos_data = []
        self.semaphore = asyncio.Semaphore(5)  # Limit concurrent clones to 5

    def read_excel_data(self):
        """
        Read data from input Excel file

        Returns:
            List of dictionaries with repo data
        """
        if not os.path.exists(self.input_file):
            raise FileNotFoundError(f"Input file '{self.input_file}' not found.")

        print(f"Reading data from {self.input_file}...")

        wb = load_workbook(self.input_file)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Find column indices
        try:
            id_col = headers.index('ID')
            timestamp_col = headers.index('TimeStamp')
            subject_col = headers.index('Subject')
            search_col = headers.index('Search Criteria')
            url_col = headers.index('github Repo URL')
        except ValueError as e:
            raise ValueError(f"Required column not found in Excel file: {e}")

        # Read data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[id_col]:  # Skip empty rows
                data.append({
                    'id': str(row[id_col]),
                    'timestamp': row[timestamp_col],
                    'subject': row[subject_col],
                    'search_criteria': row[search_col],
                    'github_url': row[url_col] if row[url_col] else '',
                    'total_lines': 0,
                    'small_files_lines': 0,
                    'grade': 0.0,
                    'status': 'pending'
                })

        print(f"Found {len(data)} repositories to analyze")
        return data

    async def clone_repo(self, repo_data):
        """
        Clone a single repository asynchronously

        Args:
            repo_data: Dictionary with repository information

        Returns:
            Updated repo_data dictionary
        """
        async with self.semaphore:
            repo_id = repo_data['id']
            github_url = repo_data['github_url']

            if not github_url or github_url.strip() == '':
                print(f"[{repo_id}] Skipping - No GitHub URL")
                repo_data['status'] = 'no_url'
                return repo_data

            # Create temp directory if it doesn't exist
            os.makedirs(self.temp_dir, exist_ok=True)

            # Create folder for this repo
            repo_folder = os.path.join(self.temp_dir, repo_id)

            # Remove existing folder if it exists
            if os.path.exists(repo_folder):
                shutil.rmtree(repo_folder)

            print(f"[{repo_id}] Cloning {github_url}...")

            # Clone the repository
            process = await asyncio.create_subprocess_exec(
                'git', 'clone', '--depth', '1', github_url, repo_folder,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )

            stdout, stderr = await process.communicate()

            if process.returncode != 0:
                error_msg = stderr.decode('utf-8', errors='ignore')
                print(f"[{repo_id}] ✗ Clone failed: {error_msg.strip()}")
                repo_data['status'] = 'clone_failed'
                return repo_data

            print(f"[{repo_id}] ✓ Clone successful")
            repo_data['status'] = 'cloned'
            repo_data['repo_folder'] = repo_folder

            return repo_data

    def count_lines_in_file(self, file_path):
        """
        Count lines in a single file

        Args:
            file_path: Path to file

        Returns:
            Number of lines in file
        """
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return sum(1 for _ in f)
        except Exception:
            return 0

    def analyze_repo(self, repo_data):
        """
        Analyze a cloned repository and count lines

        Args:
            repo_data: Dictionary with repository information

        Returns:
            Updated repo_data dictionary
        """
        if repo_data['status'] != 'cloned':
            return repo_data

        repo_folder = repo_data['repo_folder']
        repo_id = repo_data['id']

        print(f"[{repo_id}] Analyzing code...")

        total_lines = 0
        small_files_lines = 0
        file_count = 0
        small_file_count = 0

        # Walk through all files in the repository
        for root, dirs, files in os.walk(repo_folder):
            # Skip .git directory
            if '.git' in dirs:
                dirs.remove('.git')

            for file in files:
                file_path = os.path.join(root, file)

                # Count lines in file
                line_count = self.count_lines_in_file(file_path)

                if line_count > 0:
                    total_lines += line_count
                    file_count += 1

                    # Check if file has less than threshold lines
                    if line_count < self.small_file_threshold:
                        small_files_lines += line_count
                        small_file_count += 1

        # Calculate grade (percentage of code in small files)
        if total_lines > 0:
            grade = (small_files_lines / total_lines) * 100
        else:
            grade = 0.0

        repo_data['total_lines'] = total_lines
        repo_data['small_files_lines'] = small_files_lines
        repo_data['grade'] = round(grade, 2)
        repo_data['status'] = 'analyzed'

        print(f"[{repo_id}] ✓ Analysis complete: {total_lines} total lines, "
              f"{small_files_lines} lines in small files (<{self.small_file_threshold} lines), "
              f"Grade: {repo_data['grade']}%")

        return repo_data

    async def clone_all_repos(self, repos_data):
        """
        Clone all repositories in parallel

        Args:
            repos_data: List of repository data dictionaries

        Returns:
            List of updated repository data dictionaries
        """
        print(f"\n{'='*70}")
        print(f"Starting parallel repository cloning")
        print(f"{'='*70}\n")

        tasks = [self.clone_repo(repo) for repo in repos_data]
        results = await asyncio.gather(*tasks)

        return results

    def analyze_all_repos(self, repos_data):
        """
        Analyze all cloned repositories

        Args:
            repos_data: List of repository data dictionaries

        Returns:
            List of updated repository data dictionaries
        """
        print(f"\n{'='*70}")
        print(f"Starting repository analysis")
        print(f"{'='*70}\n")

        for repo_data in repos_data:
            self.analyze_repo(repo_data)

        return repos_data

    def export_to_excel(self, repos_data):
        """
        Export analyzed data to Excel file

        Args:
            repos_data: List of repository data dictionaries
        """
        print(f"\nExporting results to {self.output_file}...")

        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Repo Analysis"

        # Define headers
        headers = [
            'ID',
            'TimeStamp',
            'Subject',
            'Search Criteria',
            'github Repo URL',
            'Total Lines',
            f'Lines in Small Files (<{self.small_file_threshold})',
            'Grade (%)'
        ]
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add data rows
        for repo in repos_data:
            ws.append([
                repo['id'],
                repo['timestamp'],
                repo['subject'],
                repo['search_criteria'],
                repo['github_url'],
                repo['total_lines'] if repo['status'] == 'analyzed' else 'N/A',
                repo['small_files_lines'] if repo['status'] == 'analyzed' else 'N/A',
                repo['grade'] if repo['status'] == 'analyzed' else 'N/A'
            ])

        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # ID
        ws.column_dimensions['B'].width = 20  # TimeStamp
        ws.column_dimensions['C'].width = 50  # Subject
        ws.column_dimensions['D'].width = 30  # Search Criteria
        ws.column_dimensions['E'].width = 60  # github Repo URL
        ws.column_dimensions['F'].width = 15  # Total Lines
        ws.column_dimensions['G'].width = 25  # Lines in Small Files
        ws.column_dimensions['H'].width = 15  # Grade

        # Save the workbook
        wb.save(self.output_file)
        print(f"✓ Successfully exported to {self.output_file}")

    async def run(self, cleanup=True):
        """
        Run the complete analysis pipeline

        Args:
            cleanup: Whether to remove cloned repositories after analysis

        Returns:
            List of analyzed repository data
        """
        try:
            # Read input data
            repos_data = self.read_excel_data()

            # Clone all repositories in parallel
            repos_data = await self.clone_all_repos(repos_data)

            # Analyze all cloned repositories
            repos_data = self.analyze_all_repos(repos_data)

            # Export results to Excel
            self.export_to_excel(repos_data)

            # Print summary
            self.print_summary(repos_data)

            # Cleanup temporary files
            if cleanup:
                self.cleanup_temp_files()

            return repos_data

        except Exception as e:
            print(f"\nError: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)

    def print_summary(self, repos_data):
        """
        Print summary of analysis

        Args:
            repos_data: List of repository data dictionaries
        """
        print(f"\n{'='*70}")
        print(f"Analysis Summary")
        print(f"{'='*70}")

        total = len(repos_data)
        analyzed = sum(1 for r in repos_data if r['status'] == 'analyzed')
        failed = sum(1 for r in repos_data if r['status'] == 'clone_failed')
        no_url = sum(1 for r in repos_data if r['status'] == 'no_url')

        print(f"Total Repositories: {total}")
        print(f"Successfully Analyzed: {analyzed}")
        print(f"Clone Failed: {failed}")
        print(f"No URL: {no_url}")

        if analyzed > 0:
            total_lines_sum = sum(r['total_lines'] for r in repos_data if r['status'] == 'analyzed')
            avg_lines = total_lines_sum / analyzed
            print(f"\nAverage Lines per Repo: {int(avg_lines):,}")

    def cleanup_temp_files(self):
        """Remove temporary cloned repositories"""
        if os.path.exists(self.temp_dir):
            print(f"\nCleaning up temporary files in {self.temp_dir}...")
            shutil.rmtree(self.temp_dir)
            print(f"✓ Cleanup complete")


async def main():
    """Main function to run the analyzer"""
    parser = argparse.ArgumentParser(
        description='Analyze Repos Agent - Clone GitHub repositories and analyze code metrics',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example:
  # Analyze repositories from Output_12.xlsx
  python analyze_repos.py --input Output_12.xlsx

  # Specify custom output file
  python analyze_repos.py --input Output_12.xlsx --output custom_output.xlsx

  # Keep cloned repositories (don't cleanup)
  python analyze_repos.py --input Output_12.xlsx --no-cleanup

  # Use custom temp directory
  python analyze_repos.py --input Output_12.xlsx --temp-dir MyRepos
        """
    )

    parser.add_argument(
        '--input',
        type=str,
        default='Output_12.xlsx',
        help='Input Excel file with GitHub URLs (default: Output_12.xlsx)'
    )

    parser.add_argument(
        '--output',
        type=str,
        default='Output_23.xlsx',
        help='Output Excel file with analysis results (default: Output_23.xlsx)'
    )

    parser.add_argument(
        '--temp-dir',
        type=str,
        default='TempFiles',
        help='Directory to store cloned repositories (default: TempFiles)'
    )

    parser.add_argument(
        '--no-cleanup',
        action='store_true',
        help='Keep cloned repositories after analysis (default: cleanup)'
    )

    parser.add_argument(
        '--small-file-threshold',
        type=int,
        default=150,
        help='Maximum line count for a file to be considered "small" (default: 150)'
    )

    args = parser.parse_args()

    # Create and run analyzer
    analyzer = RepoAnalyzer(
        input_file=args.input,
        output_file=args.output,
        temp_dir=args.temp_dir,
        small_file_threshold=args.small_file_threshold
    )

    await analyzer.run(cleanup=not args.no_cleanup)


if __name__ == '__main__':
    asyncio.run(main())
