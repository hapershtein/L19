#!/usr/bin/env python3
"""
Email Drafter Agent
Creates Gmail draft messages for feedback from analyzed repositories
"""

import os
import sys
import pickle
import argparse
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

from logger_config import LoggerConfig

# Setup logger
logger = LoggerConfig.setup_logger('email_drafter')

from googleapiclient.errors import HttpError
import openpyxl
from email.mime.text import MIMEText
import base64

# If modifying these scopes, delete the file token_drafter.pickle
# Note: This uses a separate token file from gmail_agent.py because it requires
# different scopes (gmail.compose vs gmail.readonly)
SCOPES = ['https://www.googleapis.com/auth/gmail.compose']


class EmailDrafter:
    """Create Gmail draft messages from feedback messages"""

    def __init__(self, input_file='Output_34.xlsx', credentials_file='credentials.json'):
        """
        Initialize EmailDrafter

        Args:
            input_file: Excel file with feedback messages
            credentials_file: Path to OAuth credentials
        """
        self.input_file = input_file
        self.credentials_file = credentials_file
        self.token_file = 'token_drafter.pickle'  # Separate token file for compose scope
        self.service = None
        self.data = []

    def authenticate(self):
        """Authenticate with Gmail API"""
        logger.info("Starting Gmail API authentication for draft creation")
        logger.info(f"Using separate token file: {self.token_file} (for gmail.compose scope)")
        creds = None

        # Load saved credentials
        if os.path.exists(self.token_file):
            logger.debug(f"Loading existing token from {self.token_file}")
            with open(self.token_file, 'rb') as token:
                creds = pickle.load(token)

        # If no valid credentials, authenticate
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())
                except Exception as e:
                    print(f"Token refresh failed: {e}")
                    print("Re-authenticating...")
                    creds = None

            if not creds:
                if not os.path.exists(self.credentials_file):
                    print(f"Error: Credentials file '{self.credentials_file}' not found!")
                    print("Please download your OAuth 2.0 credentials from Google Cloud Console")
                    sys.exit(1)

                flow = InstalledAppFlow.from_client_secrets_file(
                    self.credentials_file, SCOPES)

                try:
                    creds = flow.run_local_server(port=0, open_browser=True)
                except Exception as e:
                    print(f"\nCouldn't open browser automatically: {e}")
                    print("\nUsing console authentication flow instead...")
                    print("Please copy the URL below and paste it into your browser,")
                    print("then paste the authorization code back here.\n")
                    creds = flow.run_console()

            # Save credentials for next run
            logger.debug(f"Saving credentials to {self.token_file}")
            with open(self.token_file, 'wb') as token:
                pickle.dump(creds, token)
            logger.info("Credentials saved for future use")

        # Build Gmail service
        self.service = build('gmail', 'v1', credentials=creds)
        logger.info("Successfully authenticated with Gmail API")
        print("✓ Successfully authenticated with Gmail API\n")

    def read_excel_data(self):
        """Read feedback messages from Excel file"""
        logger.info(f"Reading data from {self.input_file}")
        print(f"Reading data from {self.input_file}...")

        if not os.path.exists(self.input_file):
            print(f"Error: Input file '{self.input_file}' not found!")
            sys.exit(1)

        try:
            wb = openpyxl.load_workbook(self.input_file)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]

            # Find required columns
            required_columns = ['ID', 'Feedback Message']
            column_indices = {}

            for col in required_columns:
                try:
                    # Try exact match first
                    column_indices[col] = headers.index(col)
                except ValueError:
                    # Try partial match
                    for idx, header in enumerate(headers):
                        if header and col.lower() in header.lower():
                            column_indices[col] = idx
                            break
                    else:
                        print(f"Error: Required column '{col}' not found in Excel file")
                        print(f"Available columns: {headers}")
                        sys.exit(1)

            # Optional: get email recipient if available (from Subject or other fields)
            email_col = None
            subject_col = None
            for idx, header in enumerate(headers):
                if header:
                    if 'subject' in header.lower():
                        subject_col = idx
                    elif 'email' in header.lower() or 'recipient' in header.lower():
                        email_col = idx

            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[column_indices['ID']]:
                    continue

                repo_id = row[column_indices['ID']]
                feedback_message = row[column_indices['Feedback Message']]

                if not feedback_message or feedback_message == 'N/A':
                    continue

                data_row = {
                    'id': str(repo_id),
                    'feedback': str(feedback_message),
                }

                # Add optional fields
                if subject_col is not None and row[subject_col]:
                    data_row['subject'] = str(row[subject_col])
                if email_col is not None and row[email_col]:
                    data_row['recipient'] = str(row[email_col])

                self.data.append(data_row)

        except Exception as e:
            print(f"Error reading Excel file: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)

        logger.info(f"Found {len(self.data)} feedback messages to draft")
        print(f"Found {len(self.data)} feedback messages to draft\n")
        return self.data

    def create_draft(self, repo_id, feedback_message, subject_prefix=None):
        """
        Create a Gmail draft message

        Args:
            repo_id: Repository ID for subject line
            feedback_message: The feedback message content
            subject_prefix: Optional subject prefix

        Returns:
            Draft ID if successful, None otherwise
        """
        try:
            # Create subject line
            if subject_prefix:
                subject = f"{subject_prefix} - Feedback message to {repo_id}"
            else:
                subject = f"Feedback message to {repo_id}"

            # Create message
            message = MIMEText(feedback_message)
            message['subject'] = subject

            # Encode message
            raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')

            # Create draft
            draft_body = {
                'message': {
                    'raw': raw_message
                }
            }

            draft = self.service.users().drafts().create(
                userId='me',
                body=draft_body
            ).execute()

            return draft['id']

        except HttpError as error:
            logger.error(f"HttpError creating draft for {repo_id}: {error}")
            LoggerConfig.log_exception(logger, error, f"create_draft for {repo_id}")
            print(f"  ✗ Error creating draft: {error}")
            return None
        except Exception as e:
            logger.error(f"Unexpected error creating draft for {repo_id}: {e}")
            LoggerConfig.log_exception(logger, e, f"create_draft for {repo_id}")
            print(f"  ✗ Unexpected error: {e}")
            return None

    def create_all_drafts(self):
        """Create Gmail drafts for all feedback messages"""
        print(f"{'='*70}")
        print(f"Creating Gmail Draft Messages")
        print(f"{'='*70}\n")

        success_count = 0
        failed_count = 0
        results = []

        for idx, data_row in enumerate(self.data, 1):
            repo_id = data_row['id']
            feedback = data_row['feedback']
            subject_prefix = data_row.get('subject', None)

            logger.info(f"[{idx}/{len(self.data)}] Creating draft for ID: {repo_id}")
            print(f"[{idx}/{len(self.data)}] Creating draft for ID: {repo_id}")

            draft_id = self.create_draft(repo_id, feedback, subject_prefix)

            if draft_id:
                logger.info(f"Draft created successfully for {repo_id}, draft_id={draft_id}")
                print(f"  ✓ Draft created successfully (ID: {draft_id})")
                success_count += 1
                results.append({
                    'id': repo_id,
                    'draft_id': draft_id,
                    'status': 'success'
                })
            else:
                logger.error(f"Failed to create draft for {repo_id}")
                print(f"  ✗ Failed to create draft")
                failed_count += 1
                results.append({
                    'id': repo_id,
                    'draft_id': None,
                    'status': 'failed'
                })

        # Summary
        print(f"\n{'='*70}")
        print(f"Draft Creation Summary")
        print(f"{'='*70}")
        print(f"Total Feedback Messages: {len(self.data)}")
        print(f"Drafts Created: {success_count}")
        print(f"Failed: {failed_count}")
        print(f"\n✓ Drafts are available in your Gmail account under 'Drafts' folder")

        return {
            'total': len(self.data),
            'success': success_count,
            'failed': failed_count,
            'results': results
        }

    def run(self):
        """Main execution flow"""
        # Authenticate with Gmail
        self.authenticate()

        # Read feedback data
        self.read_excel_data()

        if not self.data:
            print("No feedback messages found to draft. Exiting.")
            return None

        # Create drafts
        results = self.create_all_drafts()

        return results


def main():
    session_id = LoggerConfig.get_session_id()
    logger.info("=" * 70)
    logger.info(f"Email Drafter session started - Session ID: {session_id}")
    logger.info("=" * 70)
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Create Gmail draft messages from feedback messages',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Create drafts from default file
  python email_drafter.py

  # Use custom input file
  python email_drafter.py --input feedback_repos.xlsx

  # Use custom credentials
  python email_drafter.py --credentials my_credentials.json
        """
    )

    parser.add_argument(
        '--input',
        default='Output_34.xlsx',
        help='Input Excel file with feedback messages (default: Output_34.xlsx)'
    )

    parser.add_argument(
        '--credentials',
        default='credentials.json',
        help='Path to OAuth 2.0 credentials file (default: credentials.json)'
    )

    args = parser.parse_args()

    # Create drafter
    drafter = EmailDrafter(
        input_file=args.input,
        credentials_file=args.credentials
    )

    # Run
    try:
        drafter.run()
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
