"""
Excel export functionality for Gmail data
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from logger_config import LoggerConfig

logger = LoggerConfig.setup_logger('gmail_agent')


class ExcelExporter:
    """Export email data to Excel files"""

    @staticmethod
    def export_to_excel(email_data, output_file='gmail_export.xlsx'):
        """
        Export email data to Excel file

        Args:
            email_data: List of email data dictionaries
            output_file: Output Excel file path
        """
        logger.info(f"Starting Excel export to {output_file}")

        if not email_data:
            logger.warning("No email data to export")
            print("No data to export.")
            return

        logger.info(f"Exporting {len(email_data)} emails to {output_file}")
        print(f"Exporting {len(email_data)} emails to {output_file}...")

        wb = Workbook()
        ws = wb.active
        ws.title = "Gmail Export"

        headers = ['ID', 'TimeStamp', 'Subject', 'Search Criteria', 'github Repo URL']
        ws.append(headers)

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for email in email_data:
            ws.append([
                email['id'],
                email['timestamp'],
                email['subject'],
                email['search_criteria'],
                email.get('repo_url', '')
            ])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 60

        logger.debug(f"Saving workbook to {output_file}")
        wb.save(output_file)
        logger.info(f"Successfully exported {len(email_data)} emails to {output_file}")
        print(f"Successfully exported to {output_file}")
