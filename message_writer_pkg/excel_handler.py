"""
Excel file operations for message writing
"""
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from logger_config import LoggerConfig

logger = LoggerConfig.setup_logger('message_writer')


class ExcelHandler:
    """Handle Excel file operations for message writer"""

    @staticmethod
    def read_input_file(input_file):
        """
        Read data from input Excel file

        Args:
            input_file: Path to input Excel file with grades

        Returns:
            List of dictionaries with repo data and grades
        """
        if not os.path.exists(input_file):
            logger.error(f"Input file not found: {input_file}")
            raise FileNotFoundError(f"Input file '{input_file}' not found.")

        logger.info(f"Reading data from {input_file}")
        print(f"Reading data from {input_file}...")

        logger.debug(f"Loading workbook: {input_file}")
        wb = load_workbook(input_file)
        ws = wb.active
        logger.debug(f"Workbook loaded, reading headers")

        headers = [cell.value for cell in ws[1]]

        try:
            id_col = headers.index('ID')
            timestamp_col = headers.index('TimeStamp')
            subject_col = headers.index('Subject')
            search_col = headers.index('Search Criteria')
            url_col = headers.index('github Repo URL')
            total_lines_col = headers.index('Total Lines')
            small_files_col = next(i for i, h in enumerate(headers) if 'Lines in Small Files' in str(h))
            grade_col = next(i for i, h in enumerate(headers) if 'Grade' in str(h))
        except (ValueError, StopIteration) as e:
            logger.error(f"Column parsing error: {e}")
            raise ValueError(f"Required column not found in Excel file: {e}")

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[id_col]:
                grade_val = row[grade_col]
                if isinstance(grade_val, str):
                    grade_val = grade_val.replace('%', '').strip()
                try:
                    grade = float(grade_val) if grade_val not in ['N/A', None, ''] else 0.0
                except (ValueError, TypeError):
                    grade = 0.0

                data.append({
                    'id': str(row[id_col]),
                    'timestamp': row[timestamp_col],
                    'subject': row[subject_col],
                    'search_criteria': row[search_col],
                    'github_url': row[url_col],
                    'total_lines': row[total_lines_col],
                    'small_files_lines': row[small_files_col],
                    'grade': grade,
                    'message': ''
                })

        print(f"Found {len(data)} repositories to process")
        return data

    @staticmethod
    def export_results(data, output_file):
        """
        Export data with messages to Excel file

        Args:
            data: List of repository data with messages
            output_file: Path to output Excel file
        """
        if not data:
            print("No data to export.")
            return

        print(f"\nExporting results to {output_file}...")

        wb = Workbook()
        ws = wb.active
        ws.title = "Repo Analysis with Feedback"

        headers = [
            'ID',
            'TimeStamp',
            'Subject',
            'Search Criteria',
            'github Repo URL',
            'Total Lines',
            'Lines in Small Files',
            'Grade (%)',
            'Feedback Message'
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for repo in data:
            ws.append([
                repo['id'],
                repo['timestamp'],
                repo['subject'],
                repo['search_criteria'],
                repo['github_url'],
                repo['total_lines'],
                repo['small_files_lines'],
                repo['grade'],
                repo['message']
            ])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 100

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        wb.save(output_file)
        logger.info(f"Successfully exported to {output_file}")
        print(f"✓ Successfully exported to {output_file}")
