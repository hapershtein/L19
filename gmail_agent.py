#!/usr/bin/env python3
"""
Gmail Agent - Retrieve and export emails to Excel
"""
import os
import pickle
import argparse
import re
import base64
from datetime import datetime
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dateutil import parser as date_parser
from logger_config import LoggerConfig

# Gmail API scopes
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

# Setup logger
logger = LoggerConfig.setup_logger('gmail_agent')


class GmailAgent:
    """Agent to retrieve Gmail messages and export to Excel"""

    def __init__(self, credentials_file='credentials.json'):
        """
        Initialize the Gmail Agent

        Args:
            credentials_file: Path to the OAuth 2.0 credentials JSON file
        """
        self.credentials_file = credentials_file
        self.token_file = 'token.pickle'
        self.service = None

    def authenticate(self):
        """Authenticate with Gmail API using OAuth 2.0"""
        logger.info("Starting Gmail API authentication")
        creds = None

        # Check if token.pickle exists (stores user's access and refresh tokens)
        if os.path.exists(self.token_file):
            logger.debug(f"Loading existing token from {self.token_file}")
            with open(self.token_file, 'rb') as token:
                creds = pickle.load(token)
            logger.info("Loaded existing authentication token")

        # If there are no valid credentials, let the user log in
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                logger.info("Token expired, refreshing access token")
                print("Refreshing access token...")
                creds.refresh(Request())
                logger.info("Access token refreshed successfully")
            else:
                if not os.path.exists(self.credentials_file):
                    logger.error(f"Credentials file not found: {self.credentials_file}")
                    raise FileNotFoundError(
                        f"Credentials file '{self.credentials_file}' not found. "
                        "Please download it from Google Cloud Console."
                    )

                logger.info("Starting OAuth 2.0 authentication flow")
                print("Starting OAuth 2.0 authentication flow...")
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.credentials_file, SCOPES
                )

                # Try to run local server with browser, fallback to console if it fails
                try:
                    logger.debug("Attempting browser authentication")
                    creds = flow.run_local_server(port=0, open_browser=True)
                    logger.info("Browser authentication successful")
                except Exception as e:
                    logger.warning(f"Browser authentication failed: {e}")
                    logger.info("Falling back to console authentication")
                    print(f"\nCouldn't open browser automatically: {e}")
                    print("\nUsing console authentication flow instead...")
                    print("You'll receive a URL to visit in your browser.\n")
                    creds = flow.run_console()
                    logger.info("Console authentication successful")

            # Save the credentials for the next run
            logger.debug(f"Saving credentials to {self.token_file}")
            with open(self.token_file, 'wb') as token:
                pickle.dump(creds, token)
            logger.info("Credentials saved for future use")

        self.service = build('gmail', 'v1', credentials=creds)
        logger.info("Gmail API service built successfully")
        print("Successfully authenticated with Gmail API")

    def extract_first_url(self, text):
        """
        Extract the first URL found in text

        Args:
            text: Text to search for URLs

        Returns:
            First URL found or empty string
        """
        if not text:
            return ''

        # URL pattern - matches only https://github.com URLs
        url_pattern = r'https?://github\.com[^\s<>"{}|\\^`\[\]]+'

        match = re.search(url_pattern, text)
        if match:
            url = match.group(0)
            # Clean up common trailing characters
            url = re.sub(r'[,;.\)]+$', '', url)
            return url

        return ''

    def get_message_body(self, payload):
        """
        Extract text from email message payload

        Args:
            payload: Gmail message payload

        Returns:
            Email body text
        """
        body = ''

        if 'body' in payload and 'data' in payload['body']:
            body = base64.urlsafe_b64decode(payload['body']['data']).decode('utf-8', errors='ignore')
        elif 'parts' in payload:
            for part in payload['parts']:
                if part['mimeType'] == 'text/plain':
                    if 'data' in part['body']:
                        body = base64.urlsafe_b64decode(part['body']['data']).decode('utf-8', errors='ignore')
                        break
                elif part['mimeType'] == 'text/html' and not body:
                    if 'data' in part['body']:
                        body = base64.urlsafe_b64decode(part['body']['data']).decode('utf-8', errors='ignore')
                elif 'parts' in part:
                    # Recursively check nested parts
                    body = self.get_message_body(part)
                    if body:
                        break

        return body

    def search_emails(self, query, max_results=100):
        """
        Search for emails using Gmail search syntax

        Args:
            query: Gmail search query (e.g., 'from:example@gmail.com subject:invoice')
            max_results: Maximum number of emails to retrieve

        Returns:
            List of email data dictionaries
        """
        logger.info(f"Starting email search with query: '{query}', max_results: {max_results}")

        if not self.service:
            logger.error("Gmail service not initialized - authentication required")
            raise RuntimeError("Not authenticated. Call authenticate() first.")

        try:
            print(f"Searching for emails with query: '{query}'")
            logger.debug(f"Calling Gmail API messages().list() with userId='me', query='{query}'")
            results = self.service.users().messages().list(
                userId='me',
                q=query,
                maxResults=max_results
            ).execute()

            messages = results.get('messages', [])
            logger.info(f"API returned {len(messages)} message IDs")

            if not messages:
                logger.info("No messages found matching query")
                print("No messages found.")
                return []

            print(f"Found {len(messages)} messages. Retrieving details...")
            logger.info(f"Starting to retrieve full details for {len(messages)} messages")

            email_data = []
            for i, message in enumerate(messages, 1):
                msg_id = message['id']
                logger.debug(f"Processing message {i}/{len(messages)}, ID: {msg_id}")

                # Get full message details (including body for URL extraction)
                msg = self.service.users().messages().get(
                    userId='me',
                    id=msg_id,
                    format='full'
                ).execute()

                # Extract headers
                headers = msg['payload']['headers']
                subject = next((h['value'] for h in headers if h['name'] == 'Subject'), 'No Subject')
                date_str = next((h['value'] for h in headers if h['name'] == 'Date'), '')
                from_email = next((h['value'] for h in headers if h['name'] == 'From'), '')
                logger.debug(f"Message {i}: Subject='{subject[:50]}...', From='{from_email}'")

                # Parse date
                try:
                    date_obj = date_parser.parse(date_str)
                    timestamp = date_obj.strftime('%Y-%m-%d %H:%M:%S')
                except Exception as e:
                    logger.warning(f"Failed to parse date '{date_str}': {e}")
                    timestamp = date_str

                # Extract body and find first URL
                body = self.get_message_body(msg['payload'])
                repo_url = self.extract_first_url(body)
                if repo_url:
                    logger.debug(f"Message {i}: Extracted URL: {repo_url}")
                else:
                    logger.debug(f"Message {i}: No URL found in body")

                email_data.append({
                    'id': msg_id,
                    'timestamp': timestamp,
                    'subject': subject,
                    'from': from_email,
                    'search_criteria': query,
                    'repo_url': repo_url
                })

                if i % 10 == 0:
                    print(f"  Retrieved {i}/{len(messages)} messages...")
                    logger.info(f"Progress: Retrieved {i}/{len(messages)} messages")

            logger.info(f"Successfully retrieved all {len(email_data)} emails")
            print(f"Successfully retrieved {len(email_data)} emails")
            return email_data

        except HttpError as error:
            logger.error(f"Gmail API HttpError: {error}")
            LoggerConfig.log_exception(logger, error, "search_emails")
            print(f"An error occurred: {error}")
            return []
        except Exception as error:
            logger.error(f"Unexpected error in search_emails: {error}")
            LoggerConfig.log_exception(logger, error, "search_emails")
            return []

    def export_to_excel(self, email_data, output_file='gmail_export.xlsx'):
        """
        Export email data to Excel file

        Args:
            email_data: List of email data dictionaries
            output_file: Output Excel file path
        """
        logger.info(f"Starting Excel export to {output_file}")

        if not email_data:
            logger.warning("No email data to export")
            print("No data to export.")
            return

        logger.info(f"Exporting {len(email_data)} emails to {output_file}")
        print(f"Exporting {len(email_data)} emails to {output_file}...")

        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Gmail Export"

        # Define headers
        headers = ['ID', 'TimeStamp', 'Subject', 'Search Criteria', 'github Repo URL']
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add data rows
        for email in email_data:
            ws.append([
                email['id'],
                email['timestamp'],
                email['subject'],
                email['search_criteria'],
                email.get('repo_url', '')
            ])

        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # ID
        ws.column_dimensions['B'].width = 20  # TimeStamp
        ws.column_dimensions['C'].width = 50  # Subject
        ws.column_dimensions['D'].width = 30  # Search Criteria
        ws.column_dimensions['E'].width = 60  # Repo URL

        # Save the workbook
        logger.debug(f"Saving workbook to {output_file}")
        wb.save(output_file)
        logger.info(f"Successfully exported {len(email_data)} emails to {output_file}")
        print(f"Successfully exported to {output_file}")


def main():
    """Main function to run the Gmail agent"""
    parser = argparse.ArgumentParser(
        description='Gmail Agent - Retrieve and export emails to Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Search for emails from a specific sender
  python gmail_agent.py --query "from:example@gmail.com" --output results.xlsx

  # Search for emails with specific subject
  python gmail_agent.py --query "subject:invoice" --max 50

  # Search for unread emails
  python gmail_agent.py --query "is:unread"

  # Complex search
  python gmail_agent.py --query "from:boss@company.com after:2024/01/01 has:attachment"
        """
    )

    parser.add_argument(
        '--query',
        type=str,
        required=True,
        help='Gmail search query (e.g., "from:example@gmail.com subject:invoice")'
    )

    parser.add_argument(
        '--output',
        type=str,
        default='gmail_export.xlsx',
        help='Output Excel file path (default: gmail_export.xlsx)'
    )

    parser.add_argument(
        '--max',
        type=int,
        default=100,
        help='Maximum number of emails to retrieve (default: 100)'
    )

    parser.add_argument(
        '--credentials',
        type=str,
        default='credentials.json',
        help='Path to OAuth 2.0 credentials file (default: credentials.json)'
    )

    args = parser.parse_args()

    # Create and run the agent
    try:
        session_id = LoggerConfig.get_session_id()
        logger.info(f"{'='*70}")
        logger.info(f"Gmail Agent session started - Session ID: {session_id}")
        logger.info(f"Parameters: query='{args.query}', max={args.max}, output='{args.output}'")
        logger.info(f"{'='*70}")

        agent = GmailAgent(credentials_file=args.credentials)
        agent.authenticate()

        # Search for emails
        emails = agent.search_emails(args.query, max_results=args.max)

        # Export to Excel
        if emails:
            agent.export_to_excel(emails, output_file=args.output)
            logger.info(f"Session completed successfully - {len(emails)} emails exported")
            print(f"\n✓ Complete! {len(emails)} emails exported to {args.output}")
        else:
            logger.info("Session completed - No emails found")
            print("\nNo emails found matching the search criteria.")

    except FileNotFoundError as e:
        logger.error(f"FileNotFoundError: {e}")
        print(f"\nError: {e}")
        print("\nPlease follow the setup instructions in README.md")
    except Exception as e:
        logger.error(f"Unexpected error in main: {e}")
        LoggerConfig.log_exception(logger, e, "main function")
        print(f"\nAn error occurred: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()
